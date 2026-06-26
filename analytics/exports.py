"""Excel (.xlsx) and PDF exporters for analytics datasets.

Kept dependency-light: openpyxl for Excel, reportlab for PDF. Both are imported
lazily inside the builders so the rest of the analytics app loads even if a
deployment hasn't installed them yet.
"""
from django.http import HttpResponse


def build_rows_export(fmt, filename, header, rows, title=''):
    """Return an HttpResponse for `fmt` in {'xlsx', 'pdf'}. Falls back to a
    400-style plain response for unknown formats."""
    if fmt == 'xlsx':
        return _xlsx_response(filename, header, rows, title)
    if fmt == 'pdf':
        return _pdf_response(filename, header, rows, title)
    return HttpResponse(f'Unsupported format: {fmt}', status=400)


def _xlsx_response(filename, header, rows, title):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = (title or 'Export')[:31]  # Excel sheet-name limit

    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(row)

    # Reasonable column widths.
    for i, col in enumerate(header, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(14, len(str(col)) + 2)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    wb.save(response)
    return response


def _pdf_response(filename, header, rows, title):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    import io

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=1 * cm, rightMargin=1 * cm,
                            topMargin=1 * cm, bottomMargin=1 * cm)
    styles = getSampleStyleSheet()
    elements = [Paragraph(title or filename, styles['Title']), Spacer(1, 12)]

    # Cap PDF rows so the document stays manageable.
    capped = rows[:2000]
    data = [header] + capped
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#9333ea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f3ff')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(table)
    if len(rows) > len(capped):
        elements.append(Spacer(1, 8))
        elements.append(Paragraph(
            f'Showing first {len(capped)} of {len(rows)} rows. Use CSV/Excel for the full export.',
            styles['Italic']))

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
    return response
